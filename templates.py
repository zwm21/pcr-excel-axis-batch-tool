"""
处理模板系统

每个模板定义一套独立的 Excel 处理流程。用户可在 UI 中自由切换。
模板处理器接收 App 实例以访问日志、用户设置等。
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

try:
    from wcwidth import wcswidth
    HAS_WCWIDTH = True
except ImportError:
    HAS_WCWIDTH = False

    def wcswidth(text):
        width = 0
        for ch in text:
            if '\u4e00' <= ch <= '\u9fff':
                width += 2
            else:
                width += 1
        return width


# ============================================================
#  共享工具函数（原 App 类的静态方法，提取到模块级供模板复用）
# ============================================================
MERGE_WIDTH_LIMIT = 38


def display_width(text: str) -> int:
    """计算字符串显示宽度（中文约 2 个半角字符宽度）。"""
    return wcswidth(text)


def format_sec(sec_str: str) -> str:
    """将 '1:19' 转为 '119'，'52' 转为 '052'，长度 >3 时保持原样。"""
    s = sec_str.replace(':', '')
    if len(s) <= 3:
        s = s.zfill(3)
    return s


def match_pattern(vals: list) -> bool:
    """检查 6 个值是否符合表头模式：秒数|角色|操作|操作|操作|伤害。"""
    if len(vals) != 6:
        return False
    if vals[0] != "秒数":
        return False
    if vals[1] != "角色":
        return False
    if vals[2] != "操作":
        return False
    if vals[5] != "伤害":
        return False
    ok4 = vals[3] in ("操作", "")
    ok5 = vals[4] in ("操作", "")
    return ok4 and ok5


def find_header(ws, max_row: int = 30):
    """在指定行数内查找表头，返回 (行号, 秒数所在列号) 或 None。"""
    for row in range(1, max_row + 1):
        for col in range(1, ws.max_column - 5):
            values = []
            for offset in range(6):
                cell = ws.cell(row=row, column=col + offset)
                val = cell.value
                if isinstance(val, str):
                    val = val.strip()
                else:
                    val = str(val).strip() if val is not None else ""
                values.append(val)
            if match_pattern(values):
                return row, col
    return None


def unmerge_in_rect(ws, min_row, min_col, max_row, max_col):
    """解除与指定矩形区域相交的所有合并单元格。"""
    to_unmerge = []
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= max_row and merged.max_row >= min_row and
                merged.min_col <= max_col and merged.max_col >= min_col):
            to_unmerge.append(merged)
    for rng in to_unmerge:
        ws.unmerge_cells(str(rng))


# ============================================================
#  模板元数据
# ============================================================
TEMPLATES = [
    {
        "id": "standard",
        "name": "标准处理（模板1）",
        "summary": "完整处理流程：合并单元格 + 富文本着色 + 智能行分组。"
                    "检测相同秒数的相邻行，将其合并为一行（用 \"-\" 分隔），"
                    "超出宽度限制时自动截断。",
        "description": (
            "【处理步骤】\n"
            "1. 查找表头\n"
            "   在前 30 行内定位包含「秒数｜角色｜操作｜操作｜操作｜伤害」的表头行。\n\n"
            "2. 表头处理\n"
            "   将表头行中「角色」至「伤害」共 5 列合并为一个单元格，写入「角色」。\n\n"
            "3. 逐行处理数据\n"
            "   从表头下一行开始，读取每行的「角色」和「操作」列，生成富文本：\n"
            "   \u00b7 操作为空或「连点」 -> 黑色纯角色名\n"
            "   \u00b7 操作为「AUTO」   -> 蓝色「角色名(AUTO)」\n"
            "   \u00b7 其他操作         -> 红色「角色名(操作名)」\n"
            "   然后将该行的 5 列合并为一个单元格，写入生成的富文本。\n\n"
            "4. 字体统一替换\n"
            "   将整个工作表中所有非空单元格的字体替换为「汉仪文黑-65W」，\n"
            "   同时保留原有的字号、颜色等属性。\n\n"
            "5. 格式化调整\n"
            "   取消 B1 和 C1 单元格的加粗样式；若启用了追加文字选项，\n"
            "   则在 C1 单元格末尾添加用户自定义内容。\n\n"
            "6. 智能行分组合并\n"
            "   扫描所有数据行，将相邻且「秒数」相同的行捆绑为一组：\n"
            "   · 同秒数组 → 合并为一行，用「-」连接，不添加秒数后缀\n"
            "   · 不同秒数组 → 在第一个角色名后添加秒数标记，如「(119)」\n"
            "   · 合并后总显示宽度超过限制(38个半角)时，自动截断前缀\n"
            "   · 被合并的行隐藏（保留行，不删除）\n\n"
            "7. 保存输出\n"
            "   在同一目录下生成「已抄轴_原文件名.xlsx」。"
        ),
    },
    {
        "id": "simple",
        "name": "简化处理（模板2）",
        "summary": "跳过智能行分组合并，保留每行独立输出。"
                    "仅完成基础合并、富文本着色和字体替换，"
                    "适合需要保持原有行结构的场景。",
        "description": (
            "【处理步骤】\n"
            "1. 查找表头\n"
            "   在前 30 行内定位包含「秒数｜角色｜操作｜操作｜操作｜伤害」的表头行。\n\n"
            "2. 表头处理\n"
            "   将表头行中「角色」至「伤害」共 5 列合并为一个单元格，写入「角色」。\n\n"
            "3. 逐行处理数据\n"
            "   从表头下一行开始，读取每行的「角色」和「操作」列，生成富文本：\n"
            "   \u00b7 操作为空或「连点」 -> 黑色纯角色名\n"
            "   \u00b7 操作为「AUTO」   -> 蓝色「角色名(AUTO)」\n"
            "   \u00b7 其他操作         -> 红色「角色名(操作名)」\n"
            "   然后将该行的 5 列合并为一个单元格，写入生成的富文本。\n\n"
            "4. 字体统一替换\n"
            "   将整个工作表中所有非空单元格的字体替换为「汉仪文黑-65W」，\n"
            "   同时保留原有的字号、颜色等属性。\n\n"
            "5. 格式化调整\n"
            "   取消 B1 和 C1 单元格的加粗样式；若启用了追加文字选项，\n"
            "   则在 C1 单元格末尾添加用户自定义内容。\n\n"
            "6. 保存输出\n"
            "   在同一目录下生成「已抄轴_原文件名.xlsx」。\n\n"
            "【与模板 1 的区别】\n"
            "本模板不执行「智能行分组合并」步骤，每行数据保持独立。"
            "适合不需要压缩行数、或需要保留原始行结构的场景。",
        ),
    },
]

# 默认模板
DEFAULT_TEMPLATE_ID = "standard"


# ============================================================
#  模板处理器
# ============================================================
def get_template_by_id(template_id: str) -> dict:
    """通过 ID 获取模板元数据。"""
    for tmpl in TEMPLATES:
        if tmpl["id"] == template_id:
            return tmpl
    return TEMPLATES[0]  # 回退到第一个


def process_file(filepath: str, app, template_id: str) -> str:
    """根据模板 ID 调度对应的处理函数。"""
    if template_id == "simple":
        return _process_simple(filepath, app)
    else:
        return _process_standard(filepath, app)


# ============================================================
#  模板 1：标准处理（完整流程）
# ============================================================
def _process_standard(filepath: str, app) -> str:
    """标准处理：包含智能行分组合并的完整流程。"""
    wb = load_workbook(filepath, rich_text=True)
    if "轴模板" not in wb.sheetnames:
        raise ValueError("找不到工作表 '轴模板'")
    ws = wb["轴模板"]

    # 步骤 1-2：查找并处理表头
    pos = find_header(ws, max_row=30)
    if pos is None:
        raise ValueError("前30行内未找到符合要求的表头（秒数|角色|操作|操作|操作|伤害）")
    header_row, start_col = pos
    role_col = start_col + 1
    action_col = start_col + 2
    merge_start_col = role_col
    merge_end_col = start_col + 5

    # 处理表头行
    unmerge_in_rect(ws, header_row, merge_start_col, header_row, merge_end_col)
    merge_range = f"{get_column_letter(merge_start_col)}{header_row}:{get_column_letter(merge_end_col)}{header_row}"
    ws.merge_cells(merge_range)
    header_cell = ws.cell(row=header_row, column=merge_start_col)
    header_cell.value = "角色"
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 步骤 3：逐行处理数据
    current_row = header_row + 1
    while True:
        role_value = ws.cell(row=current_row, column=role_col).value
        if role_value is None or (isinstance(role_value, str) and role_value.strip() == ""):
            break

        # 跳过宽度超过5的合并单元格
        merged = None
        for m in ws.merged_cells.ranges:
            if m.min_row <= current_row <= m.max_row and m.min_col <= role_col <= m.max_col:
                merged = m
                break
        if merged and (merged.max_col - merged.min_col + 1) > 5:
            current_row += 1
            continue

        # 读取操作文本，生成富文本
        op_value = ws.cell(row=current_row, column=action_col).value
        op_text = "" if op_value is None else str(op_value).strip()

        rt = CellRichText()
        if op_text == "" or op_text == "连点":
            fill_text = str(role_value).strip()
            rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), fill_text))
        elif op_text == "AUTO":
            fill_text = f"{role_value}(AUTO)"
            font = InlineFont(rFont='汉仪文黑-65W', color='00b0f0')
            rt.append(TextBlock(font, fill_text))
        else:
            fill_text = f"{role_value}({op_text})"
            font = InlineFont(rFont='汉仪文黑-65W', color='FF0000')
            rt.append(TextBlock(font, fill_text))

        unmerge_in_rect(ws, current_row, merge_start_col, current_row, merge_end_col)
        merge_range_data = f"{get_column_letter(merge_start_col)}{current_row}:{get_column_letter(merge_end_col)}{current_row}"
        ws.merge_cells(merge_range_data)
        data_cell = ws.cell(row=current_row, column=merge_start_col)
        data_cell.value = rt
        data_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    # 步骤 4：统一替换字体
    _replace_all_fonts(ws)

    # 步骤 5：取消 B1/C1 加粗
    _unbold_b1_c1(ws)

    # 步骤 5b：追加文字
    _apply_suffix(ws, app)

    # 步骤 6：智能行分组合并
    _merge_adjacent_rows(ws, header_row, start_col, merge_start_col)

    # 步骤 7：保存
    return _save_result(wb, filepath)


# ============================================================
#  模板 2：简化处理（不含行分组合并）
# ============================================================
def _process_simple(filepath: str, app) -> str:
    """简化处理：完成基础合并和字体替换，不进行行分组合并。"""
    wb = load_workbook(filepath, rich_text=True)
    if "轴模板" not in wb.sheetnames:
        raise ValueError("找不到工作表 '轴模板'")
    ws = wb["轴模板"]

    # 步骤 1-2：查找并处理表头
    pos = find_header(ws, max_row=30)
    if pos is None:
        raise ValueError("前30行内未找到符合要求的表头（秒数|角色|操作|操作|操作|伤害）")
    header_row, start_col = pos
    role_col = start_col + 1
    action_col = start_col + 2
    merge_start_col = role_col
    merge_end_col = start_col + 5

    # 处理表头行
    unmerge_in_rect(ws, header_row, merge_start_col, header_row, merge_end_col)
    merge_range = f"{get_column_letter(merge_start_col)}{header_row}:{get_column_letter(merge_end_col)}{header_row}"
    ws.merge_cells(merge_range)
    header_cell = ws.cell(row=header_row, column=merge_start_col)
    header_cell.value = "角色"
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 步骤 3：逐行处理数据（与模板1相同，但跳过宽合并检查以保持简洁）
    current_row = header_row + 1
    while True:
        role_value = ws.cell(row=current_row, column=role_col).value
        if role_value is None or (isinstance(role_value, str) and role_value.strip() == ""):
            break

        op_value = ws.cell(row=current_row, column=action_col).value
        op_text = "" if op_value is None else str(op_value).strip()

        rt = CellRichText()
        if op_text == "" or op_text == "连点":
            fill_text = str(role_value).strip()
            rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), fill_text))
        elif op_text == "AUTO":
            fill_text = f"{role_value}(AUTO)"
            font = InlineFont(rFont='汉仪文黑-65W', color='00b0f0')
            rt.append(TextBlock(font, fill_text))
        else:
            fill_text = f"{role_value}({op_text})"
            font = InlineFont(rFont='汉仪文黑-65W', color='FF0000')
            rt.append(TextBlock(font, fill_text))

        unmerge_in_rect(ws, current_row, merge_start_col, current_row, merge_end_col)
        merge_range_data = f"{get_column_letter(merge_start_col)}{current_row}:{get_column_letter(merge_end_col)}{current_row}"
        ws.merge_cells(merge_range_data)
        data_cell = ws.cell(row=current_row, column=merge_start_col)
        data_cell.value = rt
        data_cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    # 步骤 4：统一替换字体
    _replace_all_fonts(ws)

    # 步骤 5：取消 B1/C1 加粗 + 追加文字
    _unbold_b1_c1(ws)
    _apply_suffix(ws, app)

    # 步骤 6：保存
    return _save_result(wb, filepath)


# ============================================================
#  共享子步骤
# ============================================================
def _replace_all_fonts(ws):
    """将工作表中所有非空单元格字体替换为汉仪文黑-65W。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                old_font = cell.font
                if old_font is not None:
                    cell.font = Font(
                        name='汉仪文黑-65W',
                        size=old_font.size,
                        bold=old_font.bold,
                        italic=old_font.italic,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color=old_font.color,
                        scheme=old_font.scheme,
                        family=old_font.family,
                        charset=old_font.charset,
                    )
                else:
                    cell.font = Font(name='汉仪文黑-65W')


def _unbold_b1_c1(ws):
    """取消 B1 和 C1 单元格的加粗。"""
    for cell in [ws['B1'], ws['C1']]:
        if cell.value is not None:
            old_font = cell.font
            if old_font is not None:
                cell.font = Font(
                    name='汉仪文黑-65W',
                    size=old_font.size,
                    bold=False,
                    italic=old_font.italic,
                    underline=old_font.underline,
                    strike=old_font.strike,
                    color=old_font.color,
                    scheme=old_font.scheme,
                    family=old_font.family,
                    charset=old_font.charset,
                )
            else:
                cell.font = Font(name='汉仪文黑-65W', bold=False)


def _apply_suffix(ws, app):
    """若启用追加文字，则在 C1 末尾添加自定义内容。"""
    c1 = ws['C1']
    if app.append_text_var.get():
        suffix = app.text_suffix_var.get()
        if suffix:
            old_val = str(c1.value) if c1.value else ''
            c1.value = old_val + suffix


def _save_result(wb, filepath: str) -> str:
    """保存为「已抄轴_原文件名.xlsx」并返回新路径。"""
    import os
    dirname = os.path.dirname(filepath)
    basename = os.path.basename(filepath)
    new_name = "已抄轴_" + basename
    new_path = os.path.join(dirname, new_name)
    wb.save(new_path)
    return new_path


def _merge_adjacent_rows(ws, header_row: int, start_col: int, merge_start_col: int):
    """模板1专属：智能行分组合并（同秒相邻行合并为一行）。"""
    rows = []
    current = header_row + 1
    while current <= ws.max_row:
        cell = ws.cell(row=current, column=merge_start_col)
        if cell.value is not None:
            rows.append(current)
        current += 1

    def is_oversize_merged(row):
        for m in ws.merged_cells.ranges:
            if m.min_row <= row <= m.max_row and m.min_col <= merge_start_col <= m.max_col:
                if m.max_col - m.min_col + 1 > 5:
                    return True
        return False

    to_delete = []
    i = 0
    while i < len(rows) - 1:
        cur_row = rows[i]
        group_rows = []
        group_sec = None
        j = i + 1
        while j < len(rows):
            candidate = rows[j]
            if is_oversize_merged(candidate):
                break
            sec = str(ws.cell(row=candidate, column=start_col).value or "").strip()
            if group_sec is None:
                group_sec = sec
                group_rows.append(candidate)
            elif sec == group_sec:
                group_rows.append(candidate)
            else:
                break
            j += 1

        if not group_rows:
            i += 1
            continue
        if is_oversize_merged(cur_row):
            i += 1
            continue

        cur_cell = ws.cell(row=cur_row, column=merge_start_col)
        cur_sec = str(ws.cell(row=cur_row, column=start_col).value or "").strip()
        cur_text = _get_plain_text(cur_cell.value)
        cur_width = display_width(cur_text)

        group_plain_parts = []
        group_text_width = 0
        for idx, gr in enumerate(group_rows):
            gr_cell = ws.cell(row=gr, column=merge_start_col)
            gr_text = _get_plain_text(gr_cell.value)
            group_plain_parts.append(gr_text)
            if idx == 0:
                group_text_width += display_width(gr_text)
            else:
                group_text_width += display_width("-") + display_width(gr_text)

        if group_sec != cur_sec:
            formatted = format_sec(group_sec)
            first_gr = group_plain_parts[0]
            if "(AUTO)" in first_gr:
                modified_first = first_gr.replace("(AUTO)", f"({formatted} AUTO)")
            elif "(" in first_gr and ")" in first_gr:
                s_pos = first_gr.find("(")
                e_pos = first_gr.find(")")
                inner = first_gr[s_pos + 1:e_pos]
                modified_first = first_gr[:s_pos + 1] + formatted + " " + inner + first_gr[e_pos:]
            else:
                modified_first = first_gr + f"({formatted})"
            actual_group_width = (group_text_width
                                  - display_width(first_gr)
                                  + display_width(modified_first))
            total_width = cur_width + display_width("-") + actual_group_width
        else:
            total_width = cur_width + display_width("-") + group_text_width

        if total_width > MERGE_WIDTH_LIMIT:
            if group_sec == cur_sec:
                sep_w = display_width("-")
                best_k = -1
                for k in range(len(group_rows)):
                    subset_w = cur_width + sep_w
                    for idx in range(k + 1):
                        if idx > 0:
                            subset_w += sep_w
                        subset_w += display_width(group_plain_parts[idx])
                    if subset_w <= MERGE_WIDTH_LIMIT:
                        best_k = k
                    else:
                        break
                if best_k >= 0:
                    group_rows = group_rows[:best_k + 1]
                    group_plain_parts = group_plain_parts[:best_k + 1]
                    group_text_width = 0
                    for idx, part in enumerate(group_plain_parts):
                        if idx == 0:
                            group_text_width += display_width(part)
                        else:
                            group_text_width += sep_w + display_width(part)
                else:
                    i += 1
                    continue
            else:
                i += 1
                continue

        # 构建新富文本
        new_rt = CellRichText()
        _append_cell_rich_text(new_rt, cur_cell.value)
        new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), "-"))

        for idx, gr in enumerate(group_rows):
            gr_cell = ws.cell(row=gr, column=merge_start_col)
            if idx > 0:
                new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), "-"))
            if idx == 0 and group_sec != cur_sec:
                formatted = format_sec(group_sec)
                _append_with_sec_suffix(new_rt, gr_cell.value, formatted)
            else:
                _append_cell_rich_text(new_rt, gr_cell.value)

        cur_cell.value = new_rt
        cur_cell.alignment = Alignment(horizontal='center', vertical='center')
        to_delete.extend(group_rows)
        for _ in group_rows:
            if i + 1 < len(rows):
                rows.pop(i + 1)

    # 隐藏被合并的行
    for del_row in sorted(to_delete, reverse=True):
        if 1 <= del_row <= ws.max_row:
            ws.row_dimensions[del_row].hidden = True
            cell = ws.cell(row=del_row, column=merge_start_col)
            cell.value = None


def _get_plain_text(value) -> str:
    """从单元格值提取纯文本。"""
    if isinstance(value, CellRichText):
        return "".join(str(b) for b in value)
    return str(value) if value else ""


def _append_cell_rich_text(rt: CellRichText, value):
    """将单元格的富文本内容追加到目标 CellRichText，统一字体。"""
    def _ensure_font(block):
        if isinstance(block, TextBlock):
            if block.font is None or block.font.rFont is None:
                old_color = block.font.color if block.font else None
                new_font = InlineFont(rFont='汉仪文黑-65W', color=old_color)
                return TextBlock(new_font, block.text)
            return block
        return TextBlock(InlineFont(rFont='汉仪文黑-65W'), str(block))

    if isinstance(value, CellRichText):
        for block in value:
            rt.append(_ensure_font(block))
    else:
        rt.append(_ensure_font(str(value) if value else ""))


def _append_with_sec_suffix(rt: CellRichText, value, formatted: str):
    """追加文本，在第一个块后添加秒数标记。"""
    if isinstance(value, CellRichText):
        blocks = list(value)
        if blocks:
            first_block = blocks[0] if isinstance(blocks[0], TextBlock) else None
            if first_block:
                color_obj = first_block.font.color if first_block.font else None
                original_color = color_obj.rgb if color_obj and hasattr(color_obj, 'rgb') else None
                raw_text = str(first_block.text) if first_block.text else ""
                if "(AUTO)" in raw_text:
                    modified = raw_text.replace("(AUTO)", f"({formatted} AUTO)")
                elif "(" in raw_text and ")" in raw_text:
                    s_pos = raw_text.find("(")
                    e_pos = raw_text.find(")")
                    inner = raw_text[s_pos + 1:e_pos] if s_pos + 1 < e_pos else ""
                    modified = raw_text[:s_pos + 1] + formatted + " " + inner + raw_text[e_pos:]
                else:
                    modified = raw_text + f"({formatted})"
                rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W', color=original_color), modified))
                for b in blocks[1:]:
                    _append_cell_rich_text(rt, b)
            else:
                for b in blocks:
                    _append_cell_rich_text(rt, b)
                rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), f"({formatted})"))
        else:
            rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), _get_plain_text(value) + f"({formatted})"))
    else:
        raw = _get_plain_text(value)
        if "(AUTO)" in raw:
            modified = raw.replace("(AUTO)", f"({formatted} AUTO)")
        elif "(" in raw and ")" in raw:
            s_pos = raw.find("(")
            e_pos = raw.find(")")
            inner = raw[s_pos + 1:e_pos]
            modified = raw[:s_pos + 1] + formatted + " " + inner + raw[e_pos:]
        else:
            modified = raw + f"({formatted})"
        rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), modified))
