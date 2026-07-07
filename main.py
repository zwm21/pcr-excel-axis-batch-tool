try:
    from wcwidth import wcswidth
    HAS_WCWIDTH = True
except ImportError:
    HAS_WCWIDTH = False
    # 简单回退：粗略估算中文字符宽度为2
    def simple_width(s):
        width = 0
        for ch in s:
            if '\u4e00' <= ch <= '\u9fff':
                width += 2
            else:
                width += 1
        return width
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from ui_config import *

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# 合并后单行最大显示宽度（Excel列宽约38个半角字符）
MERGE_WIDTH_LIMIT = 38


# ============================================================
#  样式工具函数
# ============================================================
def apply_hover_style(widget, normal_bg, hover_bg, active_bg=None):
    """为非 ttk 控件绑定悬停 / 点击背景色变化。"""
    widget.bind("<Enter>", lambda e: widget.configure(background=hover_bg))
    widget.bind("<Leave>", lambda e: widget.configure(background=normal_bg))
    if active_bg:
        widget.bind("<Button-1>", lambda e: widget.configure(background=active_bg), add="+")
        widget.bind("<ButtonRelease-1>", lambda e: (
            widget.configure(background=hover_bg)
            if widget.winfo_containing(e.x_root, e.y_root) is widget
            else widget.configure(background=normal_bg)
        ), add="+")


def _make_font(size=FONT_SIZE_NORMAL, bold=False, family=FONT_FAMILY):
    """快捷创建字体元组，供 tkinter 控件使用。"""
    weight = "bold" if bold else "normal"
    return (family, size, weight)


def _padded(inner):
    """返回带外边距的填充值（2 × GRID）。"""
    return inner + PAD_CARD


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("轴模板批量处理工具")
        self.root.geometry(f"{WIN_DEFAULT_WIDTH}x{WIN_DEFAULT_HEIGHT}")
        self.root.minsize(WIN_MIN_WIDTH, WIN_MIN_HEIGHT)
        self.root.configure(bg=COLOR_BG)

        self.files = []       # 存储文件路径
        self.last_dir = None  # 记住上次打开的目录

        self._configure_ttk_style()
        self._build_header()
        self._build_file_section()       # row=1
        self._build_option_bar()         # row=2
        self._build_log_section()        # row=3

        # 窗口缩放权重
        self.root.grid_rowconfigure(1, weight=1)   # 文件列表卡片伸缩
        self.root.grid_rowconfigure(3, weight=1)   # 日志区卡片伸缩
        self.root.grid_columnconfigure(0, weight=1)

    # ----------------------------------------------------------
    #  ttk 主题配置
    # ----------------------------------------------------------
    def _configure_ttk_style(self):
        style = ttk.Style()
        style.theme_use("clam")

        # 通用卡片外框
        style.configure("Card.TFrame", background=COLOR_CARD_BG,
                        relief="solid", borderwidth=1)

        # 标签
        style.configure("Section.TLabel",
                        font=_make_font(FONT_SIZE_SMALL, bold=True),
                        foreground=COLOR_TEXT_SECONDARY,
                        background=COLOR_CARD_BG)

        style.configure("Title.TLabel",
                        font=_make_font(FONT_SIZE_HEADING, bold=True),
                        foreground=COLOR_TEXT,
                        background=COLOR_BG)

        # 按钮样式
        style.configure("Primary.TButton",
                        font=_make_font(FONT_SIZE_NORMAL),
                        padding=(PAD_CARD, GAP_MD),
                        background=COLOR_PRIMARY,
                        foreground="#FFFFFF",
                        borderwidth=0,
                        focuscolor="none")
        style.map("Primary.TButton",
                  background=[("active", COLOR_PRIMARY_HOVER),
                              ("pressed", COLOR_PRIMARY_ACTIVE)],
                  foreground=[("active", "#FFFFFF")])

        style.configure("Accent.TButton",
                        font=_make_font(FONT_SIZE_NORMAL, bold=True),
                        padding=(PAD_CARD * 2, GAP_MD),
                        background=COLOR_ACCENT,
                        foreground="#FFFFFF",
                        borderwidth=0,
                        focuscolor="none")
        style.map("Accent.TButton",
                  background=[("active", COLOR_ACCENT_HOVER),
                              ("pressed", COLOR_ACCENT_ACTIVE)],
                  foreground=[("active", "#FFFFFF")])

        style.configure("Danger.TButton",
                        font=_make_font(FONT_SIZE_NORMAL),
                        padding=(PAD_CARD, GAP_MD),
                        background="#E8E8E8",
                        borderwidth=0,
                        focuscolor="none")
        style.map("Danger.TButton",
                  background=[("active", "#D8D8D8"),
                              ("pressed", "#C8C8C8")])

    # ----------------------------------------------------------
    #  顶部标题栏
    # ----------------------------------------------------------
    def _build_header(self):
        frame = tk.Frame(self.root, bg=COLOR_BG)
        frame.grid(row=0, column=0, sticky="ew",
                   padx=PAD_PAGE, pady=(PAD_PAGE, GAP_LG))
        frame.grid_columnconfigure(0, weight=1)

        ttk.Label(frame, text="轴模板批量处理工具",
                  style="Title.TLabel").pack(side=tk.LEFT)

    # ----------------------------------------------------------
    #  文件列表卡片
    # ----------------------------------------------------------
    def _build_file_section(self):
        card = tk.Frame(self.root, bg=COLOR_CARD_BG,
                        highlightbackground=COLOR_CARD_BORDER,
                        highlightthickness=1)
        card.grid(row=1, column=0, sticky="nsew",
                  padx=PAD_PAGE, pady=(0, GAP_LG))
        card.grid_rowconfigure(1, weight=1)
        card.grid_columnconfigure(0, weight=1)

        # 标题行
        title_row = tk.Frame(card, bg=COLOR_CARD_BG)
        title_row.grid(row=0, column=0, sticky="ew",
                       padx=PAD_CARD, pady=(PAD_TITLE_Y, GAP_MD))

        ttk.Label(title_row, text="待处理文件列表",
                  style="Section.TLabel").pack(side=tk.LEFT)
        self._file_count_label = ttk.Label(
            title_row, text="（0 个文件）",
            font=_make_font(FONT_SIZE_SMALL),
            foreground=COLOR_TEXT_LIGHT, background=COLOR_CARD_BG)
        self._file_count_label.pack(side=tk.LEFT, padx=(GAP_MD, 0))

        # 列表框
        list_frame = tk.Frame(card, bg=COLOR_CARD_BORDER)
        list_frame.grid(row=1, column=0, sticky="nsew",
                        padx=PAD_CARD, pady=(0, GAP_MD))
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)

        self.listbox = tk.Listbox(
            list_frame,
            selectmode=tk.EXTENDED,
            height=LISTBOX_HEIGHT,
            bg=COLOR_LISTBOX_BG,
            fg=COLOR_TEXT,
            font=_make_font(FONT_SIZE_NORMAL),
            selectbackground=COLOR_LISTBOX_SELECT,
            selectforeground=COLOR_LISTBOX_SELECT_TEXT,
            activestyle="none",
            borderwidth=0,
            highlightthickness=0,
            relief="flat",
        )
        self.listbox.grid(row=0, column=0, sticky="nsew",
                          padx=1, pady=1)

        # 按钮行
        btn_row = tk.Frame(card, bg=COLOR_CARD_BG)
        btn_row.grid(row=2, column=0, sticky="ew",
                     padx=PAD_CARD, pady=(0, PAD_CARD_Y))

        self.add_files_btn = ttk.Button(
            btn_row, text="＋ 添加文件", style="Primary.TButton",
            command=self.add_files)
        self.add_files_btn.pack(side=tk.LEFT, padx=(0, GAP_SM))

        self.add_folder_btn = ttk.Button(
            btn_row, text=" 添加文件夹", style="Primary.TButton",
            command=self.add_folder)
        self.add_folder_btn.pack(side=tk.LEFT, padx=GAP_SM)

        self.clear_btn = ttk.Button(
            btn_row, text="清空列表", style="Danger.TButton",
            command=self.clear_list)
        self.clear_btn.pack(side=tk.LEFT, padx=GAP_SM)

        self._op_buttons = (self.add_files_btn, self.add_folder_btn,
                            self.clear_btn, None)  # start_btn later

    # ----------------------------------------------------------
    #  选项栏 + 开始按钮
    # ----------------------------------------------------------
    def _build_option_bar(self):
        card = tk.Frame(self.root, bg=COLOR_CARD_BG,
                        highlightbackground=COLOR_CARD_BORDER,
                        highlightthickness=1)
        card.grid(row=2, column=0, sticky="ew",
                  padx=PAD_PAGE, pady=(0, GAP_LG))
        card.grid_columnconfigure(1, weight=1)

        # 复选框
        self.append_text_var = tk.BooleanVar(value=True)
        cb = tk.Checkbutton(
            card, text="在轴标题末尾追加文字",
            variable=self.append_text_var,
            bg=COLOR_CARD_BG, fg=COLOR_TEXT,
            font=_make_font(FONT_SIZE_NORMAL),
            activebackground=COLOR_CARD_BG,
            activeforeground=COLOR_TEXT,
            selectcolor=COLOR_CARD_BG)
        cb.grid(row=0, column=0, sticky="w",
                padx=(PAD_CARD, GAP_MD), pady=PAD_CARD_Y)

        # 输入框（带小标签）
        lbl = tk.Label(card, text="内容：", bg=COLOR_CARD_BG,
                       fg=COLOR_TEXT_SECONDARY,
                       font=_make_font(FONT_SIZE_NORMAL))
        lbl.grid(row=0, column=1, sticky="e",
                 padx=(0, GAP_SM), pady=PAD_CARD_Y)

        self.text_suffix_var = tk.StringVar(value=" by 筱娅")
        self.suffix_entry = tk.Entry(
            card, textvariable=self.text_suffix_var,
            width=ENTRY_WIDTH,
            font=_make_font(FONT_SIZE_NORMAL),
            bg="#FFFFFF", fg=COLOR_TEXT,
            insertbackground=COLOR_TEXT,
            relief="solid", borderwidth=1,
            highlightbackground=COLOR_CARD_BORDER,
            highlightthickness=0)
        self.suffix_entry.grid(row=0, column=2, sticky="w",
                                padx=(0, GAP_LG), pady=PAD_CARD_Y)

        # 开始按钮（右侧）
        self.start_btn = ttk.Button(
            card, text="▶ 开始处理", style="Accent.TButton",
            command=self.start_processing)
        self.start_btn.grid(row=0, column=3, sticky="e",
                            padx=(0, PAD_CARD), pady=PAD_CARD_Y)

        # 补齐 _op_buttons
        self._op_buttons = (self.add_files_btn, self.add_folder_btn,
                            self.clear_btn, self.start_btn)

    # ----------------------------------------------------------
    #  日志区域卡片
    # ----------------------------------------------------------
    def _build_log_section(self):
        card = tk.Frame(self.root, bg=COLOR_CARD_BG,
                        highlightbackground=COLOR_CARD_BORDER,
                        highlightthickness=1)
        card.grid(row=3, column=0, sticky="nsew",
                  padx=PAD_PAGE, pady=(0, PAD_PAGE))
        card.grid_rowconfigure(1, weight=1)
        card.grid_columnconfigure(0, weight=1)

        # 标题行
        title_row = tk.Frame(card, bg=COLOR_CARD_BG)
        title_row.grid(row=0, column=0, sticky="ew",
                       padx=PAD_CARD, pady=(PAD_TITLE_Y, GAP_MD))

        ttk.Label(title_row, text="处理日志",
                  style="Section.TLabel").pack(side=tk.LEFT)

        # 日志文本框
        log_frame = tk.Frame(card, bg=COLOR_LOG_BORDER)
        log_frame.grid(row=1, column=0, sticky="nsew",
                       padx=PAD_CARD, pady=(0, PAD_CARD_Y))
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            width=80, height=LOG_HEIGHT,
            state=tk.DISABLED,
            bg=COLOR_LOG_BG,
            fg=COLOR_TEXT,
            font=_make_font(FONT_SIZE_NORMAL, family=FONT_MONO),
            insertbackground=COLOR_TEXT,
            borderwidth=0,
            highlightthickness=0,
            relief="flat",
            wrap=tk.WORD,
        )
        self.log_text.grid(row=0, column=0, sticky="nsew",
                           padx=1, pady=1)

    # ============================================================
    #  业务逻辑（以下代码未做任何修改，仅保持原样）
    # ============================================================
    def add_files(self):
        """添加选中的excel文件"""
        paths = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=self.last_dir
        )
        if paths:
            self.last_dir = os.path.dirname(paths[0])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.listbox.insert(tk.END, p)
        self._update_file_count()

    def add_folder(self):
        """添加文件夹中的所有excel文件"""
        folder = filedialog.askdirectory(title="选择文件夹", initialdir=self.last_dir)
        if folder:
            self.last_dir = folder
        for root, dirs, files in os.walk(folder):
            for f in files:
                if f.lower().endswith(".xlsx"):
                    full = os.path.join(root, f)
                    if full not in self.files:
                        self.files.append(full)
                        self.listbox.insert(tk.END, full)
        self._update_file_count()

    def _update_file_count(self):
        """更新文件计数标签。"""
        cnt = len(self.files)
        self._file_count_label.config(text=f"（{cnt} 个文件）")
        if cnt > 0:
            self._file_count_label.config(foreground=COLOR_PRIMARY)
        else:
            self._file_count_label.config(foreground=COLOR_TEXT_LIGHT)

    def clear_list(self):
        """清空文件列表"""
        self.files.clear()
        self.listbox.delete(0, tk.END)
        self._update_file_count()

    def log(self, msg):
        """向日志区域添加一条消息"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _set_buttons_state(self, state):
        """统一设置所有操作按钮的启用/禁用状态"""
        for btn in self._op_buttons:
            btn.config(state=state)

    def start_processing(self):
        """启动处理线程"""
        if not self.files:
            messagebox.showwarning("警告", "没有待处理的文件，请先添加文件或文件夹。")
            return
        self._set_buttons_state(tk.DISABLED)
        self.log("开始处理...")
        thread = threading.Thread(target=self.process_all, daemon=True)
        thread.start()

    def process_all(self):
        """依次处理每个文件"""
        success_count = 0
        fail_count = 0
        for idx, filepath in enumerate(self.files, 1):
            self.log(f"[{idx}/{len(self.files)}] 正在处理: {os.path.basename(filepath)}")
            try:
                out_path = self.process_file(filepath)
                self.log(f"✓ 处理成功 -> {out_path}")
                success_count += 1
            except Exception as e:
                self.log(f"✗ 处理失败: {str(e)}")
                fail_count += 1
        self.log(f"处理完成！成功: {success_count}, 失败: {fail_count}")
        self.root.after(0, lambda: self._set_buttons_state(tk.NORMAL))

    def process_file(self, filepath):
        """处理单个文件：查找表头、合并表头和数据行、保存新文件"""
        wb = load_workbook(filepath, rich_text=True)
        if "轴模板" not in wb.sheetnames:
            raise ValueError("找不到工作表 '轴模板'")
        ws = wb["轴模板"]

        # 查找表头位置（返回秒数所在行和列）
        pos = self.find_header(ws, max_row=30)
        if pos is None:
            raise ValueError("前30行内未找到符合要求的表头（秒数|角色|操作|操作|操作|伤害）")
        header_row, start_col = pos   # start_col 是"秒数"所在列
        role_col = start_col + 1      # "角色"所在列
        action_col = start_col + 2    # 第一个"操作"所在列
        # 合并区域：从角色列到伤害列，共5列
        merge_start_col = role_col
        merge_end_col = start_col + 5   # 伤害列

        # 1. 处理表头行
        self.unmerge_in_rect(ws, header_row, merge_start_col, header_row, merge_end_col)
        merge_range = f"{get_column_letter(merge_start_col)}{header_row}:{get_column_letter(merge_end_col)}{header_row}"
        ws.merge_cells(merge_range)
        header_cell = ws.cell(row=header_row, column=merge_start_col)
        header_cell.value = "角色"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 2. 处理数据行（从表头的下一行开始）
        current_row = header_row + 1
        while True:
            # 检查当前行的角色列是否为空（表示数据结束）
            role_value = ws.cell(row=current_row, column=role_col).value
            if role_value is None or (isinstance(role_value, str) and role_value.strip() == ""):
                break

            # 在处理数据行循环内部，读取 role_value 之后，添加以下跳过逻辑：
            merged = None
            for m in ws.merged_cells.ranges:
                if m.min_row <= current_row <= m.max_row and m.min_col <= role_col <= m.max_col:
                    merged = m
                    break
            if merged and (merged.max_col - merged.min_col + 1) > 5:
                current_row += 1
                continue

            # 读取操作文本（第一个操作列）
            op_value = ws.cell(row=current_row, column=action_col).value
            if op_value is None:
                op_text = ""
            else:
                op_text = str(op_value).strip()

            # 创建富文本对象
            rt = CellRichText()
            if op_text == "" or op_text == "连点":
                fill_text = str(role_value).strip()
                rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), fill_text))
            elif op_text == "AUTO":
                fill_text = f"{role_value}(AUTO)"
                font = InlineFont(rFont='汉仪文黑-65W', color='00b0f0')
                rt.append(TextBlock(font, fill_text))
            else:
                fill_text = f"{role_value}({op_text})"
                font = InlineFont(rFont='汉仪文黑-65W', color='FF0000')
                rt.append(TextBlock(font, fill_text))

            # 解除该行的合并区域
            self.unmerge_in_rect(ws, current_row, merge_start_col, current_row, merge_end_col)
            # 合并5格
            merge_range_data = f"{get_column_letter(merge_start_col)}{current_row}:{get_column_letter(merge_end_col)}{current_row}"
            ws.merge_cells(merge_range_data)
            data_cell = ws.cell(row=current_row, column=merge_start_col)
            data_cell.value = rt
            data_cell.alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        # --- 修改字体：仅改变字体名称，保留其他属性 ---
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    old_font = cell.font
                    if old_font is not None:
                        cell.font = Font(
                            name='汉仪文黑-65W',
                            size=old_font.size,
                            bold=old_font.bold,
                            italic=old_font.italic,
                            underline=old_font.underline,
                            strike=old_font.strike,
                            color=old_font.color,
                            scheme=old_font.scheme,
                            family=old_font.family,
                            charset=old_font.charset
                        )
                    else:
                        cell.font = Font(name='汉仪文黑-65W')

        # 取消 B1 和 C1 的加粗
        for cell in [ws['B1'], ws['C1']]:
            if cell.value is not None:
                old_font = cell.font
                if old_font is not None:
                    cell.font = Font(
                        name='汉仪文黑-65W',
                        size=old_font.size,
                        bold=False,
                        italic=old_font.italic,
                        underline=old_font.underline,
                        strike=old_font.strike,
                        color=old_font.color,
                        scheme=old_font.scheme,
                        family=old_font.family,
                        charset=old_font.charset
                    )
                else:
                    cell.font = Font(name='汉仪文黑-65W', bold=False)

        # 若启用追加文字，则在 C1 末尾添加用户自定义内容
        c1 = ws['C1']
        if self.append_text_var.get():
            suffix = self.text_suffix_var.get()
            if suffix:
                old_val = str(c1.value) if c1.value else ''
                c1.value = old_val + suffix

        # ========== 合并符合条件的相邻行（完整规则） ==========
        rows = []
        current = header_row + 1
        while current <= ws.max_row:
            cell = ws.cell(row=current, column=merge_start_col)
            if cell.value is not None:
                rows.append(current)
            current += 1

        def is_oversize_merged(row):
            for m in ws.merged_cells.ranges:
                if m.min_row <= row <= m.max_row and m.min_col <= merge_start_col <= m.max_col:
                    if m.max_col - m.min_col + 1 > 5:
                        return True
            return False

        to_delete = []

        i = 0
        while i < len(rows) - 1:
            cur_row = rows[i]

            group_rows = []
            group_sec = None
            j = i + 1
            while j < len(rows):
                candidate = rows[j]
                if is_oversize_merged(candidate):
                    break
                sec = str(ws.cell(row=candidate, column=start_col).value or "").strip()
                if group_sec is None:
                    group_sec = sec
                    group_rows.append(candidate)
                elif sec == group_sec:
                    group_rows.append(candidate)
                else:
                    break
                j += 1

            if not group_rows:
                i += 1
                continue

            if is_oversize_merged(cur_row):
                i += 1
                continue

            cur_cell = ws.cell(row=cur_row, column=merge_start_col)
            cur_sec = str(ws.cell(row=cur_row, column=start_col).value or "").strip()

            def get_plain_text(value):
                if isinstance(value, CellRichText):
                    return "".join(str(b) for b in value)
                return str(value) if value else ""

            group_text_width = 0
            group_plain_parts = []
            for idx, gr in enumerate(group_rows):
                gr_cell = ws.cell(row=gr, column=merge_start_col)
                gr_text = get_plain_text(gr_cell.value)
                group_plain_parts.append(gr_text)
                if idx == 0:
                    group_text_width += self.display_width(gr_text)
                else:
                    group_text_width += self.display_width("-") + self.display_width(gr_text)

            cur_text = get_plain_text(cur_cell.value)
            cur_width = self.display_width(cur_text)

            if group_sec != cur_sec:
                formatted = self.format_sec(group_sec)
                first_gr_text = group_plain_parts[0]
                if "(AUTO)" in first_gr_text:
                    modified_first = first_gr_text.replace("(AUTO)", f"({formatted} AUTO)")
                elif "(" in first_gr_text and ")" in first_gr_text:
                    start = first_gr_text.find("(")
                    end = first_gr_text.find(")")
                    inner = first_gr_text[start+1:end]
                    modified_first = first_gr_text[:start+1] + formatted + " " + inner + first_gr_text[end:]
                else:
                    modified_first = first_gr_text + f"({formatted})"
                actual_group_width = (group_text_width
                                      - self.display_width(first_gr_text)
                                      + self.display_width(modified_first))
                total_width = cur_width + self.display_width("-") + actual_group_width
            else:
                total_width = cur_width + self.display_width("-") + group_text_width

            if total_width > MERGE_WIDTH_LIMIT:
                if group_sec == cur_sec:
                    sep_w = self.display_width("-")
                    best_k = -1
                    for k in range(len(group_rows)):
                        subset_w = cur_width + sep_w
                        for idx in range(k + 1):
                            if idx > 0:
                                subset_w += sep_w
                            subset_w += self.display_width(group_plain_parts[idx])
                        if subset_w <= MERGE_WIDTH_LIMIT:
                            best_k = k
                        else:
                            break
                    if best_k >= 0:
                        group_rows = group_rows[:best_k + 1]
                        group_plain_parts = group_plain_parts[:best_k + 1]
                        group_text_width = 0
                        for idx, part in enumerate(group_plain_parts):
                            if idx == 0:
                                group_text_width += self.display_width(part)
                            else:
                                group_text_width += sep_w + self.display_width(part)
                        total_width = cur_width + sep_w + group_text_width
                    else:
                        i += 1
                        continue
                else:
                    i += 1
                    continue

            # ---------- 开始构建新富文本 ----------
            new_rt = CellRichText()

            def ensure_font(block):
                if isinstance(block, TextBlock):
                    if block.font is None or block.font.rFont is None:
                        old_color = block.font.color if block.font else None
                        new_font = InlineFont(rFont='汉仪文黑-65W', color=old_color)
                        return TextBlock(new_font, block.text)
                    else:
                        return block
                else:
                    return TextBlock(InlineFont(rFont='汉仪文黑-65W'), str(block))

            if isinstance(cur_cell.value, CellRichText):
                for block in cur_cell.value:
                    new_rt.append(ensure_font(block))
            else:
                new_rt.append(ensure_font(str(cur_cell.value) if cur_cell.value else ""))

            new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), "-"))

            for idx, gr in enumerate(group_rows):
                gr_cell = ws.cell(row=gr, column=merge_start_col)
                if idx > 0:
                    new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), "-"))

                if idx == 0 and group_sec != cur_sec:
                    formatted = self.format_sec(group_sec)
                    if isinstance(gr_cell.value, CellRichText):
                        blocks = list(gr_cell.value)
                        if blocks:
                            first_block = blocks[0] if isinstance(blocks[0], TextBlock) else None
                            if first_block:
                                color_obj = first_block.font.color if first_block.font else None
                                original_color = color_obj.rgb if color_obj and hasattr(color_obj, 'rgb') else None
                                raw_text = str(first_block.text) if first_block.text else ""
                                if "(AUTO)" in raw_text:
                                    modified = raw_text.replace("(AUTO)", f"({formatted} AUTO)")
                                elif "(" in raw_text and ")" in raw_text:
                                    start = raw_text.find("(")
                                    end = raw_text.find(")")
                                    inner = raw_text[start+1:end] if start+1 < end else ""
                                    modified = raw_text[:start+1] + formatted + " " + inner + raw_text[end:]
                                else:
                                    modified = raw_text + f"({formatted})"
                                new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W', color=original_color), modified))
                                for b in blocks[1:]:
                                    new_rt.append(ensure_font(b))
                            else:
                                for b in blocks:
                                    new_rt.append(ensure_font(b))
                                new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), f"({formatted})"))
                        else:
                            new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), get_plain_text(gr_cell.value) + f"({formatted})"))
                    else:
                        raw = get_plain_text(gr_cell.value)
                        if "(AUTO)" in raw:
                            modified = raw.replace("(AUTO)", f"({formatted} AUTO)")
                        elif "(" in raw and ")" in raw:
                            start = raw.find("(")
                            end = raw.find(")")
                            inner = raw[start+1:end]
                            modified = raw[:start+1] + formatted + " " + inner + raw[end:]
                        else:
                            modified = raw + f"({formatted})"
                        new_rt.append(TextBlock(InlineFont(rFont='汉仪文黑-65W'), modified))
                else:
                    if isinstance(gr_cell.value, CellRichText):
                        for block in gr_cell.value:
                            new_rt.append(ensure_font(block))
                    else:
                        new_rt.append(ensure_font(str(gr_cell.value) if gr_cell.value else ""))

            cur_cell.value = new_rt
            cur_cell.alignment = Alignment(horizontal='center', vertical='center')

            to_delete.extend(group_rows)
            for _ in group_rows:
                if i+1 < len(rows):
                    rows.pop(i+1)

        for del_row in sorted(to_delete, reverse=True):
            if 1 <= del_row <= ws.max_row:
                ws.row_dimensions[del_row].hidden = True
                cell = ws.cell(row=del_row, column=merge_start_col)
                cell.value = None

        dirname = os.path.dirname(filepath)
        basename = os.path.basename(filepath)
        new_name = "已抄轴_" + basename
        new_path = os.path.join(dirname, new_name)

        wb.save(new_path)
        return new_path

    def find_header(self, ws, max_row=30):
        """在指定行数内查找表头，返回(行号, 秒数所在列号)"""
        for row in range(1, max_row + 1):
            for col in range(1, ws.max_column - 5):
                values = []
                for offset in range(6):
                    cell = ws.cell(row=row, column=col + offset)
                    val = cell.value
                    if isinstance(val, str):
                        val = val.strip()
                    else:
                        val = str(val).strip() if val is not None else ""
                    values.append(val)
                if self.match_pattern(values):
                    return row, col
        return None

    @staticmethod
    def match_pattern(vals):
        """检查6个值是否符合表头模式"""
        if len(vals) != 6:
            return False
        if vals[0] != "秒数":
            return False
        if vals[1] != "角色":
            return False
        if vals[2] != "操作":
            return False
        if vals[5] != "伤害":
            return False
        ok4 = vals[3] in ("操作", "")
        ok5 = vals[4] in ("操作", "")
        return ok4 and ok5

    @staticmethod
    def unmerge_in_rect(ws, min_row, min_col, max_row, max_col):
        """解除与指定矩形区域相交的所有合并单元格"""
        to_unmerge = []
        for merged in ws.merged_cells.ranges:
            if (merged.min_row <= max_row and merged.max_row >= min_row and
                merged.min_col <= max_col and merged.max_col >= min_col):
                to_unmerge.append(merged)
        for rng in to_unmerge:
            ws.unmerge_cells(str(rng))

    @staticmethod
    def display_width(text):
        if HAS_WCWIDTH:
            return wcswidth(text)
        else:
            return simple_width(text)

    @staticmethod
    def format_sec(sec_str):
        """将 '1:19' 转为 '119'，'52' 转为 '052'，长度 >3 时保持原样"""
        s = sec_str.replace(':', '')
        if len(s) <= 3:
            s = s.zfill(3)
        return s


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
